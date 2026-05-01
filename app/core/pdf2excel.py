#  Copyright (c) 2017-2026 null. All rights reserved.
"""PDF转Excel模块，支持结构化表格和纯文字两种转换模式"""
import pdfplumber
import tabula
from openpyxl import Workbook

from app.utils.file_utils import get_file_path


def _save_excel_file(workbook: Workbook, excel_filename: str) -> str:
    """
    保存Excel文件

    :param workbook: Workbook对象
    :param excel_filename: 输出的Excel文件名
    :return: Excel文件完整路径
    """
    excel_path = get_file_path(excel_filename)
    workbook.save(excel_path)
    return excel_path


def convert_pdf_table_to_excel(pdf_filename: str, excel_filename: str) -> str:
    """
    PDF表格转Excel（结构化表格）

    :param pdf_filename: 上传的PDF文件名
    :param excel_filename: 输出的Excel文件名
    :return: Excel文件完整路径
    :raises RuntimeError: 转换过程失败或未检测到表格时抛出
    """
    pdf_path = get_file_path(pdf_filename)

    try:
        tables = tabula.read_pdf(pdf_path, pages="all", multiple_tables=True)
        if not tables:
            raise RuntimeError("未在PDF中检测到结构化表格")

        wb = Workbook()
        wb.remove(wb.active)

        for i, table in enumerate(tables):
            ws = wb.create_sheet(title=f"表格{i + 1}")
            for row in table.values.tolist():
                ws.append(row)

        return _save_excel_file(wb, excel_filename)
    except RuntimeError:
        raise
    except Exception as e:
        raise RuntimeError(f"PDF表格转Excel失败：{str(e)}")


def convert_pdf_text_to_excel(pdf_filename: str, excel_filename: str) -> str:
    """
    PDF纯文字转Excel（无表格场景）

    :param pdf_filename: 上传的PDF文件名
    :param excel_filename: 输出的Excel文件名
    :return: Excel文件完整路径
    :raises RuntimeError: 转换过程失败时抛出
    """
    pdf_path = get_file_path(pdf_filename)

    try:
        with pdfplumber.open(pdf_path) as pdf:
            wb = Workbook()
            ws = wb.active
            ws.title = "PDF文字内容"

            for page_num, page in enumerate(pdf.pages, 1):
                text = page.extract_text()
                if text:
                    ws.append([f"==== 第{page_num}页 ===="])
                    for line in text.split("\n"):
                        ws.append([line])
                    ws.append([""])

        return _save_excel_file(wb, excel_filename)
    except Exception as e:
        raise RuntimeError(f"PDF文字转Excel失败：{str(e)}")
